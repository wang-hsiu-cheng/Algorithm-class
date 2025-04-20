import numpy as np
import openpyxl
import matplotlib.pyplot as plt

myWorkbook = openpyxl.load_workbook('result.xlsx')
result_data = myWorkbook['hw06']

data_rows = result_data.max_row
data_columns = result_data.max_column
algorithm_time = [[0 for i in range(data_columns-1)] for j in range(data_rows-1)]
data_numbers = [0 for i in range(data_columns-1)]
algorithm_name = [0 for i in range(data_rows-1)]

# Define the input size (n)
n = np.arange(10, 1400)  # n from 1 to 100

O_n = n / 1000000000
O_log_n = np.log2(n) / 1000000000
O_n_log_n = (n) * np.log2(n) / 1000000000
O_n2 = n**2 / 1000000000
O_n3 = abs(n**3) / 1000000000

for i in range(2, data_rows+1):
    for j in range(2, data_columns+1):
        algorithm_time[i-2][j-2] = result_data.cell(row=i, column=j).value
for i in range(2, data_columns+1):
    data_numbers[i-2] = result_data.cell(row=1, column=i).value
for i in range(2, data_rows+1):
    algorithm_name[i-2] = result_data.cell(row=i, column=1).value

# set graph style
plt.figure(figsize=(8,6))
# plt.plot(t, reference, color="C8", label="O(n)", linewidth=3)
plt.plot(data_numbers, algorithm_time[0], color="gray", label=algorithm_name[0], linestyle="-", linewidth=3)
plt.plot(data_numbers, algorithm_time[1], color="orange", label=algorithm_name[1], linestyle="-", linewidth=3)
plt.plot(data_numbers, algorithm_time[2], color="blue", label=algorithm_name[2], linestyle="-", linewidth=3)
plt.plot(data_numbers, algorithm_time[3], color="pink", label=algorithm_name[3], linestyle="-", linewidth=3)
plt.plot(data_numbers, algorithm_time[4], color="C5", label=algorithm_name[4], linestyle="-", linewidth=3)
plt.plot(data_numbers, algorithm_time[5], color="C6", label=algorithm_name[5], linestyle="-", linewidth=3)
# plt.plot(n, O_n, color="C0", label=r"$O(n)$", linestyle="--", linewidth=3)
# plt.plot(n, O_log_n, color="C1", label=r"$O(log n)$", linestyle="--", linewidth=3)
plt.plot(n, O_n_log_n, color="C2", label=r"$O(n log n)$", linestyle="--", linewidth=3)
# plt.plot(n, O_n2, color="C3", label=r"$O(n^2)$", linestyle="--", linewidth=3)
# plt.plot(n, O_n3, color="C4", label=r"$O(n^3)$", linestyle="--", linewidth=3)
plt.title('Max Subarray Prob Avg Time (log-log scale)', fontsize=14)
# plt.title('Search Algorithm Worst Case Time (log-log scale)')
plt.xlabel('Number of data', fontsize=14)
plt.ylabel('Time (sec)', fontsize=14)
plt.tick_params(axis='both', labelsize=14)  # 'both' applies to x and y axes
plt.yscale('log')
plt.xscale('log')
plt.grid(True)
plt.legend(fontsize=14)
plt.show()