import numpy as np
import openpyxl
import matplotlib.pyplot as plt

myWorkbook = openpyxl.load_workbook('result.xlsx')
result_data = myWorkbook['sheet1']

data_rows = result_data.max_row
data_columns = result_data.max_column
sort_time = [[0 for i in range(data_columns-1)] for j in range(data_rows-1)]
data_numbers = [0 for i in range(data_columns-1)]
sorting_name = [0 for i in range(data_rows-1)]

for i in range(2, data_rows+1):
    for j in range(2, data_columns+1):
        sort_time[i-2][j-2] = result_data.cell(row=i, column=j).value
for i in range(2, data_columns+1):
    data_numbers[i-2] = result_data.cell(row=1, column=i).value
for i in range(2, data_rows+1):
    sorting_name[i-2] = result_data.cell(row=i, column=1).value

# set graph style
plt.figure(figsize=(8,6))
plt.plot(data_numbers, sort_time[0], color="gray", label=sorting_name[0], linewidth=3)
plt.plot(data_numbers, sort_time[1], color="orange", label=sorting_name[1], linewidth=3)
plt.plot(data_numbers, sort_time[2], color="blue", label=sorting_name[2], linewidth=3)
plt.plot(data_numbers, sort_time[3], color="pink", label=sorting_name[3], linewidth=3)
plt.title('hw01 result')
plt.xlabel('Number of data')
plt.ylabel('Time (s)')
plt.yscale('log')
plt.xscale('log')
plt.grid(True)
plt.legend()
plt.show()