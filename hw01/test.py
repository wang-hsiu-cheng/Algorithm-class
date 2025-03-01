import numpy as np
import openpyxl
import matplotlib.pyplot as plt

myWorkbook = openpyxl.load_workbook('result.xlsx')
result_data = myWorkbook['sheet1']

data_rows = result_data.max_row
data_columns = result_data.max_column
sort_time = [[0 for i in range(data_columns-1)] for j in range(data_rows-1)]
data_numbers = []
sorting_name = []

for i in range(2, data_rows+1):
    for j in range(2, data_columns):
        print(result_data.cell(row=i, column=j).value)
for i in range(2, data_columns+1):
    print(result_data.cell(row=1, column=i).value)
for i in range(2, data_rows+1):
    print(result_data.cell(row=i, column=1).value)